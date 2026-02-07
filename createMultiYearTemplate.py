#!/usr/bin/env python3
"""
Create Excel template with separate sheets for each year
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

def create_multi_year_template(json_path, output_path):
    # Load events from JSON
    with open(json_path, 'r') as f:
        data = json.load(f)
    
    events = data.get('events', [])
    print(f"📖 Loaded {len(events)} events from {json_path}")
    
    # Group events by year
    events_by_year = {}
    for event in events:
        try:
            year = event['date'].split('-')[0]
            if year not in events_by_year:
                events_by_year[year] = []
            events_by_year[year].append(event)
        except:
            print(f"⚠️  Skipping event with invalid date: {event.get('title', 'Unknown')}")
    
    print(f"📅 Found events in years: {sorted(events_by_year.keys())}")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Instructions sheet
    ws_instructions = wb.create_sheet('Instructions', 0)
    
    instructions_data = [
        ['Field', 'Description', 'Example'],
        ['event_title', 'Name of the event *REQUIRED', 'Introduction to R Workshop'],
        ['date', 'Date in YYYY-MM-DD format *REQUIRED', '2025-03-15'],
        ['time', 'Time in HH:MM 24-hour format *REQUIRED', '18:00'],
        ['venue', 'Location of the event *REQUIRED', 'Carslaw Building Room 157'],
        ['type', 'Must be: academic, social, or industry *REQUIRED', 'academic'],
        ['description', 'Brief description *REQUIRED', 'Learn Python fundamentals...'],
        ['collaborators', 'Comma-separated societies *REQUIRED', 'SUDATA, SUMS, EconSoc'],
        ['catering', 'Food/drinks provided *REQUIRED', 'Pizza and drinks'],
        ['signup_link', 'URL to sign-up form *REQUIRED', 'https://forms.google.com/...']
    ]
    
    for row_idx, row_data in enumerate(instructions_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_instructions.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='00F0FF', end_color='00F0FF', fill_type='solid')
    
    # Set column widths for instructions
    ws_instructions.column_dimensions['A'].width = 20
    ws_instructions.column_dimensions['B'].width = 50
    ws_instructions.column_dimensions['C'].width = 40
    
    # Headers for event sheets
    headers = ['event_title', 'date', 'time', 'venue', 'type', 'description', 'collaborators', 'catering', 'signup_link']
    
    # Create a sheet for each year
    for year in sorted(events_by_year.keys()):
        sheet_name = f'Events {year}'
        ws_events = wb.create_sheet(sheet_name)
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws_events.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='00F0FF', end_color='00F0FF', fill_type='solid')
        
        # Add events for this year
        year_events = sorted(events_by_year[year], key=lambda x: (x['date'], x['time']))
        for row_idx, event in enumerate(year_events, 2):
            ws_events.cell(row=row_idx, column=1, value=event.get('title', ''))
            ws_events.cell(row=row_idx, column=2, value=event.get('date', ''))
            ws_events.cell(row=row_idx, column=3, value=event.get('time', ''))
            ws_events.cell(row=row_idx, column=4, value=event.get('venue', ''))
            ws_events.cell(row=row_idx, column=5, value=event.get('type', ''))
            ws_events.cell(row=row_idx, column=6, value=event.get('description', ''))
            
            # Handle collaborators
            collaborators = event.get('collaborators', [])
            if isinstance(collaborators, list):
                collaborators_str = ', '.join(collaborators)
            else:
                collaborators_str = str(collaborators)
            ws_events.cell(row=row_idx, column=7, value=collaborators_str)
            
            ws_events.cell(row=row_idx, column=8, value=event.get('catering', 'None'))
            ws_events.cell(row=row_idx, column=9, value=event.get('signupLink', ''))
        
        # Set column widths
        ws_events.column_dimensions['A'].width = 40  # event_title
        ws_events.column_dimensions['B'].width = 12  # date
        ws_events.column_dimensions['C'].width = 10  # time
        ws_events.column_dimensions['D'].width = 30  # venue
        ws_events.column_dimensions['E'].width = 12  # type
        ws_events.column_dimensions['F'].width = 60  # description
        ws_events.column_dimensions['G'].width = 30  # collaborators
        ws_events.column_dimensions['H'].width = 20  # catering
        ws_events.column_dimensions['I'].width = 50  # signup_link
        
        print(f'✅ Created sheet "{sheet_name}" with {len(year_events)} events')
    
    # Save workbook
    wb.save(output_path)
    
    print(f'\n✅ Created {output_path}')
    print(f'📄 Sheet 1: Instructions')
    for year in sorted(events_by_year.keys()):
        print(f'📄 Sheet: Events {year} ({len(events_by_year[year])} events)')
    print("\nDirectors can:")
    print("1. Edit events in each year's sheet")
    print("2. Add new sheets for new years (e.g., 'Events 2027')")
    print("3. Run: python3 convertExcelToJson.py events_template.xlsx src/data/events.json")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('\n📘 Usage: python3 createMultiYearTemplate.py <input-json> [output-excel]')
        print('\n📝 Example:')
        print('  python3 createMultiYearTemplate.py events_copy.json events_template.xlsx\n')
        sys.exit(1)
    
    json_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'events_template.xlsx'
    
    create_multi_year_template(json_path, output_path)
