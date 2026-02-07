#!/usr/bin/env python3
"""
Convert Excel file to events.json
Supports multiple sheets for different years (e.g., "Events 2025", "Events 2026")
Excel format:
- Row 1: Headers (event_title, date, time, venue, type, description, collaborators, catering, signup_link)
- Row 2+: Actual event data
"""
import json
import sys
from openpyxl import load_workbook
from datetime import datetime
import re

def convert_excel_to_json(excel_path, output_path):
    try:
        print(f'📖 Reading Excel file: {excel_path}')
        
        # Load workbook
        wb = load_workbook(excel_path)
        
        # Find all event sheets (exclude Instructions sheet)
        event_sheets = [sheet for sheet in wb.sheetnames if sheet.lower() != 'instructions']
        
        print(f'📄 Found {len(event_sheets)} event sheet(s): {event_sheets}')
        
        all_events = []
        
        # Process each sheet
        for sheet_name in event_sheets:
            ws = wb[sheet_name]
            print(f'\n📋 Processing "{sheet_name}"...')
            
            # Read headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value.strip() if cell.value else '')
            
            print(f'🏷️  Headers: {headers}')
            
            # Validate required headers
            required = ['event_title', 'date', 'time', 'venue', 'type', 'description', 'collaborators', 'catering', 'signup_link']
            missing = [h for h in required if h not in headers]
            if missing:
                print(f'⚠️  Warning: Missing headers in {sheet_name}: {missing}')
                continue
            
            # Read data rows
            sheet_events = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Skip empty rows
                if not any(row):
                    continue
                
                # Create dict from headers and values
                row_data = dict(zip(headers, row))
                
                # Extract event title
                event_title = str(row_data.get('event_title', '')).strip()
                if not event_title:
                    print(f'⚠️  Skipping row {row_idx} (no event title)')
                    continue
                
                # Parse collaborators
                collaborators = ['SUDATA']
                if row_data.get('collaborators'):
                    collab_str = str(row_data['collaborators']).strip()
                    if collab_str:
                        collaborators = [c.strip() for c in collab_str.split(',') if c.strip()]
                
                # Format date
                date_value = row_data.get('date', '')
                if isinstance(date_value, datetime):
                    formatted_date = date_value.strftime('%Y-%m-%d')
                elif date_value:
                    date_str = str(date_value).strip()
                    # Try to parse various date formats
                    try:
                        if '/' in date_str:
                            parts = date_str.split('/')
                            if len(parts) == 3:
                                month, day, year = parts
                                formatted_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                            else:
                                formatted_date = date_str
                        elif '-' in date_str and len(date_str) == 10:
                            formatted_date = date_str
                        else:
                            formatted_date = date_str
                    except:
                        formatted_date = date_str
                else:
                    formatted_date = 'TBA'
                
                # Format time
                time_value = row_data.get('time', '')
                if isinstance(time_value, datetime):
                    formatted_time = time_value.strftime('%H:%M')
                elif isinstance(time_value, float):
                    # Excel time as fraction of day
                    total_minutes = int(time_value * 24 * 60)
                    hours = total_minutes // 60
                    minutes = total_minutes % 60
                    formatted_time = f"{hours:02d}:{minutes:02d}"
                elif time_value:
                    time_str = str(time_value).strip()
                    if ':' in time_str:
                        parts = time_str.split(':')
                        formatted_time = f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
                    else:
                        formatted_time = time_str
                else:
                    formatted_time = '00:00'
                
                # Validate type
                valid_types = ['academic', 'social', 'industry']
                event_type = str(row_data.get('type', 'social')).lower().strip()
                if event_type not in valid_types:
                    print(f'⚠️  Warning: Invalid type "{event_type}" for "{event_title}". Using "social".')
                    event_type = 'social'
                
                # Create event object
                event = {
                    'title': event_title,
                    'date': formatted_date,
                    'time': formatted_time,
                    'venue': str(row_data.get('venue', 'TBA')).strip() if row_data.get('venue') else 'TBA',
                    'type': event_type,
                    'description': str(row_data.get('description', '')).strip() if row_data.get('description') else '',
                    'collaborators': collaborators if collaborators else ['SUDATA'],
                    'catering': str(row_data.get('catering', 'None')).strip() if row_data.get('catering') else 'None',
                    'signupLink': str(row_data.get('signup_link', 'https://forms.google.com/example')).strip() if row_data.get('signup_link') else 'https://forms.google.com/example',
                    'attendees': 0
                }
                
                sheet_events.append(event)
            
            print(f'✅ Loaded {len(sheet_events)} events from "{sheet_name}"')
            all_events.extend(sheet_events)
        
        # Sort all events by date
        all_events.sort(key=lambda x: (x['date'], x['time']))
        
        # Assign IDs after sorting
        for idx, event in enumerate(all_events, 1):
            event['id'] = f"event_{str(idx).zfill(3)}"
        
        # Create output JSON
        output = {'events': all_events}
        
        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        
        print(f'\n✅ Successfully converted {len(all_events)} total events from {len(event_sheets)} sheet(s)')
        print(f'📁 Saved to: {output_path}')
        
        return len(all_events)
        
    except Exception as error:
        print(f'❌ Error: {error}')
        import traceback
        traceback.print_exc()
        raise

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('\n📘 Usage: python3 convertExcelToJson.py <excel-file> [output-file]')
        print('\n📝 Example:')
        print('  python3 convertExcelToJson.py events_template.xlsx src/data/events.json\n')
        print('💡 Tip: Create sheets named "Events 2025", "Events 2026", etc.')
        print('   All sheets (except "Instructions") will be processed.\n')
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'events.json'
    
    convert_excel_to_json(excel_path, output_path)
