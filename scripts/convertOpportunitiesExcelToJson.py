#!/usr/bin/env python3
"""
Convert Excel file to opportunities.json
Excel format (single "Opportunities" sheet):
  - Row 1: Headers
  - Row 2+: Opportunity data

Columns:
  opportunity_title  | Title of the opportunity
  sponsor            | Sponsor company name (must match a current sponsor)
  sponsor_logo       | Path to sponsor logo, e.g. /sponsors/current-sponsors/imc-trading.webp
  sponsor_tier       | Display order: 1=Industry Partner, 2=Technical, 3=Recruitment/Sponsor (optional, default 99)
  type               | Internship | Program | Full-Time & Internship | Work Experience | Scholarship | Other
  deadline           | YYYY-MM-DD, or leave blank for rolling/no deadline
  status             | open | closed
  description        | Brief description of the opportunity
  application_link   | Full URL to the application page

Usage:
  python3 convertOpportunitiesExcelToJson.py opportunities_template.xlsx src/data/opportunities.json
"""
import json
import sys
from openpyxl import load_workbook
from datetime import datetime

VALID_STATUSES = ['open', 'closed']
VALID_TYPES = ['Internship', 'Program', 'Full-Time & Internship', 'Work Experience', 'Scholarship', 'Other']
REQUIRED_HEADERS = [
    'opportunity_title', 'sponsor', 'sponsor_logo', 'type',
    'deadline', 'status', 'description', 'application_link'
]


def parse_date(value):
    """Return YYYY-MM-DD string or None for rolling/blank deadlines."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    date_str = str(value).strip()
    if not date_str or date_str.lower() in ('none', 'rolling', 'tba', 'n/a', ''):
        return None
    # Accept YYYY-MM-DD directly
    if len(date_str) == 10 and date_str[4] == '-':
        return date_str
    # Try DD/MM/YYYY or MM/DD/YYYY
    if '/' in date_str:
        parts = date_str.split('/')
        if len(parts) == 3:
            # Assume DD/MM/YYYY
            try:
                return datetime.strptime(date_str, '%d/%m/%Y').strftime('%Y-%m-%d')
            except ValueError:
                pass
            # Fallback to MM/DD/YYYY
            try:
                return datetime.strptime(date_str, '%m/%d/%Y').strftime('%Y-%m-%d')
            except ValueError:
                pass
    return date_str  # Return as-is if we can't parse


def convert_excel_to_json(excel_path, output_path):
    try:
        print(f'📖 Reading Excel file: {excel_path}')

        wb = load_workbook(excel_path)

        # Find the Opportunities sheet (skip Instructions)
        opp_sheets = [s for s in wb.sheetnames if s.lower() != 'instructions']

        if not opp_sheets:
            print('❌ No data sheets found (only "Instructions" sheet exists).')
            sys.exit(1)

        print(f'📄 Found sheet(s): {opp_sheets}')

        all_opportunities = []

        for sheet_name in opp_sheets:
            ws = wb[sheet_name]
            print(f'\n📋 Processing "{sheet_name}"...')

            # Read headers from row 1
            headers = [cell.value.strip() if cell.value else '' for cell in ws[1]]
            print(f'🏷️  Headers: {headers}')

            # Validate required headers
            missing = [h for h in REQUIRED_HEADERS if h not in headers]
            if missing:
                print(f'⚠️  Missing required headers in "{sheet_name}": {missing}')
                print('    Skipping this sheet.')
                continue

            sheet_opps = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue  # Skip blank rows

                row_data = dict(zip(headers, row))

                # --- opportunity_title (required) ---
                title = str(row_data.get('opportunity_title', '')).strip()
                if not title:
                    print(f'⚠️  Skipping row {row_idx}: no opportunity_title')
                    continue

                # --- sponsor (required) ---
                sponsor = str(row_data.get('sponsor', '')).strip()
                if not sponsor:
                    print(f'⚠️  Skipping row {row_idx} "{title}": no sponsor')
                    continue

                # --- sponsor_logo ---
                sponsor_logo = str(row_data.get('sponsor_logo', '')).strip() if row_data.get('sponsor_logo') else ''

                # --- sponsor_tier (optional, controls display order) ---
                tier_raw = row_data.get('sponsor_tier')
                try:
                    sponsor_tier = int(tier_raw) if tier_raw is not None and str(tier_raw).strip() != '' else 99
                except (ValueError, TypeError):
                    sponsor_tier = 99

                # --- type ---
                opp_type = str(row_data.get('type', 'Other')).strip()
                if opp_type not in VALID_TYPES:
                    print(f'⚠️  "{title}": type "{opp_type}" not in {VALID_TYPES}. Defaulting to "Other".')
                    opp_type = 'Other'

                # --- deadline ---
                deadline = parse_date(row_data.get('deadline'))

                # --- status ---
                status = str(row_data.get('status', 'open')).lower().strip()
                if status not in VALID_STATUSES:
                    print(f'⚠️  "{title}": status "{status}" not in {VALID_STATUSES}. Defaulting to "open".')
                    status = 'open'

                # --- description ---
                description = str(row_data.get('description', '')).strip() if row_data.get('description') else ''

                # --- application_link ---
                app_link = str(row_data.get('application_link', '')).strip() if row_data.get('application_link') else ''

                opp = {
                    'sponsor': sponsor,
                    'sponsorTier': sponsor_tier,
                    'sponsorLogo': sponsor_logo,
                    'title': title,
                    'type': opp_type,
                    'deadline': deadline,
                    'status': status,
                    'description': description,
                    'applicationLink': app_link,
                }

                sheet_opps.append(opp)

            print(f'✅ Loaded {len(sheet_opps)} opportunities from "{sheet_name}"')
            all_opportunities.extend(sheet_opps)

        # Sort: open first, then closed; within each group sort by deadline (None/rolling last)
        def sort_key(o):
            status_order = 0 if o['status'] == 'open' else 1
            deadline = o['deadline'] or 'ZZZZ'  # Push None to end
            return (status_order, deadline)

        all_opportunities.sort(key=sort_key)

        # Assign IDs after sorting
        for idx, opp in enumerate(all_opportunities, 1):
            opp['id'] = f'opp_{str(idx).zfill(3)}'

        output = {'opportunities': all_opportunities}

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)

        print(f'\n✅ Successfully converted {len(all_opportunities)} opportunities')
        print(f'📁 Saved to: {output_path}')

        return len(all_opportunities)

    except Exception as error:
        print(f'❌ Error: {error}')
        import traceback
        traceback.print_exc()
        raise


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('\n📘 Usage: python3 convertOpportunitiesExcelToJson.py <excel-file> [output-file]')
        print('\n📝 Example:')
        print('  python3 convertOpportunitiesExcelToJson.py src/data/opportunities_template.xlsx src/data/opportunities.json\n')
        print('💡 Fill in the "Opportunities" sheet, then run this script.')
        print('   All sheets except "Instructions" will be processed.\n')
        sys.exit(1)

    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'opportunities.json'

    convert_excel_to_json(excel_path, output_path)
