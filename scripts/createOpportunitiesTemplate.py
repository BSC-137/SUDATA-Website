#!/usr/bin/env python3
"""
Generate opportunities_template.xlsx from the current opportunities.json
(mirrors createMultiYearTemplate.py for events)

Usage:
  python3 createOpportunitiesTemplate.py src/data/opportunities.json src/data/opportunities_template.xlsx

After generating:
  1. Open the Excel file and add/edit rows in the "Opportunities" sheet
  2. Run convertOpportunitiesExcelToJson.py to push changes back to JSON
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

NEON = '00F0FF'   # SUDATA cyan header colour
GREY = 'D9D9D9'  # Light grey for alternating rows

HEADERS = [
    'opportunity_title',
    'sponsor',
    'sponsor_logo',
    'sponsor_tier',
    'type',
    'deadline',
    'status',
    'description',
    'application_link',
]

COLUMN_WIDTHS = {
    'A': 45,   # opportunity_title
    'B': 20,   # sponsor
    'C': 45,   # sponsor_logo
    'D': 14,   # sponsor_tier
    'E': 25,   # type
    'F': 14,   # deadline
    'G': 10,   # status
    'H': 70,   # description
    'I': 55,   # application_link
}

INSTRUCTIONS = [
    ['Field',             'Description',                                                                        'Example / Allowed Values'],
    ['opportunity_title', 'Title of the opportunity  *REQUIRED',                                                'Summer Internship Program'],
    ['sponsor',           'Sponsor company name  *REQUIRED',                                                    'IMC Trading'],
    ['sponsor_logo',      'Path to logo in /public/sponsors/current-sponsors/  *REQUIRED',                      '/sponsors/current-sponsors/imc-trading.webp'],
    ['sponsor_tier',      'Controls the display order of sponsor groups (optional). Lower = shown first.',      '1 = Industry Partner  |  2 = Technical  |  3 = Recruitment/Sponsor  |  blank = last'],
    ['type',              'Category  *REQUIRED — must be one of the allowed values',                             'Internship | Program | Full-Time & Internship | Work Experience | Scholarship | Other'],
    ['deadline',          'Application deadline in YYYY-MM-DD format. Leave blank for rolling / no deadline.',  '2026-04-30'],
    ['status',            'open  or  closed  *REQUIRED',                                                        'open'],
    ['description',       'Short paragraph describing the opportunity  *REQUIRED',                               'IMC Trading is offering internship positions for...'],
    ['application_link',  'Full URL to the application page  *REQUIRED',                                        'https://bit.ly/4cdblbA'],
]


def create_opportunities_template(json_path, output_path):
    # Load existing JSON
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    opportunities = data.get('opportunities', [])
    print(f'📖 Loaded {len(opportunities)} opportunities from {json_path}')

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ── Instructions sheet ────────────────────────────────────────────────────
    ws_instr = wb.create_sheet('Instructions', 0)

    for row_idx, row_data in enumerate(INSTRUCTIONS, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_instr.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=NEON, end_color=NEON, fill_type='solid')

    ws_instr.column_dimensions['A'].width = 22
    ws_instr.column_dimensions['B'].width = 55
    ws_instr.column_dimensions['C'].width = 65
    ws_instr.row_dimensions[1].height = 20

    # ── Opportunities sheet ───────────────────────────────────────────────────
    ws = wb.create_sheet('Opportunities')

    # Header row
    header_fill = PatternFill(start_color=NEON, end_color=NEON, fill_type='solid')
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Data rows from JSON
    for row_idx, opp in enumerate(opportunities, 2):
        # Alternate row fill for readability
        row_fill = PatternFill(start_color=GREY, end_color=GREY, fill_type='solid') if row_idx % 2 == 0 else None

        tier = opp.get('sponsorTier')
        values = [
            opp.get('title', ''),
            opp.get('sponsor', ''),
            opp.get('sponsorLogo', ''),
            tier if tier is not None else '',   # blank if not set
            opp.get('type', ''),
            opp.get('deadline', ''),            # None becomes empty string in Excel
            opp.get('status', 'open'),
            opp.get('description', ''),
            opp.get('applicationLink', ''),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else '')
            cell.alignment = Alignment(wrap_text=False, vertical='top')
            if row_fill:
                cell.fill = row_fill

    # Column widths
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze the header row
    ws.freeze_panes = 'A2'

    wb.save(output_path)

    print(f'\n✅ Created: {output_path}')
    print('📄 Sheets:')
    print('   • Instructions  — field guide for the spons team')
    print(f'  • Opportunities — {len(opportunities)} existing rows pre-filled')
    print('\nWorkflow for the spons team:')
    print('  1. Open opportunities_template.xlsx and add/edit rows in the "Opportunities" sheet')
    print('  2. Save the file')
    print('  3. Run:  python3 scripts/convertOpportunitiesExcelToJson.py \\')
    print('                   src/data/opportunities_template.xlsx \\')
    print('                   src/data/opportunities.json')
    print('  4. The website will pick up the updated JSON on next build/deploy\n')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('\n📘 Usage: python3 createOpportunitiesTemplate.py <input-json> [output-excel]')
        print('\n📝 Example:')
        print('  python3 createOpportunitiesTemplate.py src/data/opportunities.json src/data/opportunities_template.xlsx\n')
        sys.exit(1)

    json_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'opportunities_template.xlsx'

    create_opportunities_template(json_path, output_path)
